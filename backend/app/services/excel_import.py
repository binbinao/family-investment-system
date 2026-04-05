"""Excel import service for holdings and transactions."""

import io
import logging
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.holding import Holding
from app.models.transaction import Transaction

logger = logging.getLogger(__name__)

# 表头别名 → 内部字段名（第一行表头匹配任一别名即可）
HOLDING_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "symbol": ("标识代码", "标的代码", "代码"),
    "name": ("名称", "标的名称"),
    "asset_type": ("资产类型",),
    "quantity": ("数量", "持股数量", "持有份额", "份额", "余额", "张数"),
    "cost_price": ("单位成本", "成本价", "每股成本", "成本"),
    "latest_price": ("最新价", "现价", "最新价格", "单位净值"),
    "account": ("账户",),
}

TRANSACTION_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "symbol": ("标识代码", "标的代码", "代码"),
    "tx_type": ("交易类型",),
    "quantity": ("数量",),
    "price": ("价格", "单价", "成交价", "净值"),
    "tx_date": ("日期", "交易日期"),
    "fee": ("手续费",),
}

HOLDING_REQUIRED_FIELDS = ("symbol", "name", "asset_type", "quantity", "cost_price")

# 旧版模板列序（无「最新价」列）：账户在第 6 列
LEGACY_HOLDING_COLS: dict[str, int | None] = {
    "symbol": 0,
    "name": 1,
    "asset_type": 2,
    "quantity": 3,
    "cost_price": 4,
    "latest_price": None,
    "account": 5,
}

TRANSACTION_REQUIRED_FIELDS = ("symbol", "tx_type", "quantity", "price", "tx_date")

LEGACY_TRANSACTION_COLS: dict[str, int | None] = {
    "symbol": 0,
    "tx_type": 1,
    "quantity": 2,
    "price": 3,
    "tx_date": 4,
    "fee": 5,
}

VALID_ASSET_TYPES = {"股票", "基金", "债券", "现金", "其他"}
VALID_TX_TYPES = {"买入", "卖出", "现金分红", "红利再投资"}


def _norm_header(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    for suffix in ("（可选）", "(可选)"):
        if s.endswith(suffix):
            s = s[: -len(suffix)].strip()
    return s


def _resolve_columns(
    header_row: tuple,
    field_aliases: dict[str, tuple[str, ...]],
    required_fields: tuple[str, ...],
    legacy_cols: dict[str, int | None],
) -> dict[str, int | None]:
    headers = list(header_row)
    norm = [_norm_header(h) for h in headers]
    cols: dict[str, int | None] = {k: None for k in field_aliases}
    for field, aliases in field_aliases.items():
        for i, nh in enumerate(norm):
            if nh in aliases:
                cols[field] = i
                break
    if all(cols.get(f) is not None for f in required_fields):
        return cols
    return legacy_cols


def _cell(row: tuple, cols: dict[str, int | None], key: str):
    i = cols.get(key)
    if i is None or i >= len(row):
        return None
    return row[i]


def create_holding_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "持仓导入"
    headers = [
        "标识代码",
        "名称",
        "资产类型",
        "数量",
        "单位成本",
        "最新价(可选)",
        "账户(可选)",
    ]
    ws.append(headers)
    ws.append(["600519", "贵州茅台", "股票", 100, 1800.00, 1850.00, "张三-华泰"])
    ws.append(["005827", "易方达蓝筹精选", "基金", 5000, 2.50, 2.55, ""])
    ws.append(["019666", "示例国债", "债券", 100, 101.20, 100.50, ""])
    ws.append(["CNY-活期", "工商银行活期", "现金", 50000, 1, 1, ""])
    ws.append(["ALT-01", "其他资产示例", "其他", 10, 100.00, "", ""])

    ws_help = wb.create_sheet("填写说明", 1)
    help_lines = [
        "【持仓导入】各「资产类型」填写规则（与网页记账一致）",
        "",
        "股票：数量为股数；单位成本=元/股；最新价可选=当前市价（元/股）。",
        "基金：数量为份额；单位成本=元/份（建仓净值）；最新价可选=当前净值（元/份）。",
        "债券：数量一般为张数；单位成本=净价（元）；最新价可选=估值净价。",
        "现金：数量=账户余额（元）；单位成本请填 1（留空时导入将自动按 1）；最新价可填 1 或留空。",
        "其他：按自定义单位填写数量与单位成本。",
        "",
        "请勿修改第一列表头名称；可增加行，不要合并数据区单元格。",
    ]
    for i, line in enumerate(help_lines, start=1):
        ws_help.cell(row=i, column=1, value=line)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_transaction_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "交易导入"
    headers = [
        "标识代码",
        "交易类型",
        "数量",
        "价格",
        "日期",
        "手续费(可选)",
    ]
    ws.append(headers)
    ws.append(["600519", "买入", 100, 1800.00, "2026-01-15", 5.00])
    ws.append(["005827", "买入", 2000, 1.25, "2026-01-16", ""])
    ws.append(["019666", "卖出", 20, 100.80, "2026-02-01", ""])
    ws.append(["CNY-活期", "买入", 10000, 1, "2026-02-10", ""])
    ws.append(["CNY-活期", "卖出", 3000, 1, "2026-02-15", ""])
    ws.append(["600519", "现金分红", 1, 320.00, "2026-06-20", ""])
    ws.append(["005827", "红利再投资", 50, 1.28, "2026-06-25", ""])

    ws_help = wb.create_sheet("填写说明", 1)
    help_lines = [
        "【交易导入】与网页「记录交易」语义一致；须已存在同「标识代码」的持仓。",
        "",
        "股票：买入/卖出填股数与成交价（元/股）；现金分红可填 数量=1、价格=分红总额（元）；红利再投资填新增份额与净值。",
        "基金：买入/卖出填份额与净值（元/份）；红利再投资填份额与确认净值。",
        "债券：填张数与净价（元）。",
        "现金：仅支持「买入」（存入）与「卖出」（取出）；数量=金额（元），价格固定填 1。",
        "",
        "日期格式：YYYY-MM-DD。手续费可为空。",
    ]
    for i, line in enumerate(help_lines, start=1):
        ws_help.cell(row=i, column=1, value=line)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_decimal(value, field_name: str) -> tuple[Decimal | None, str | None]:
    if value is None:
        return None, f"{field_name}不能为空"
    try:
        d = Decimal(str(value))
        if d <= 0:
            return None, f"{field_name}必须大于0"
        return d, None
    except (InvalidOperation, ValueError):
        return None, f"{field_name}格式错误: {value}"


def _parse_decimal_optional(
    value, field_name: str
) -> tuple[Decimal | None, str | None]:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        return None, None
    try:
        d = Decimal(str(value))
        if d <= 0:
            return None, f"{field_name}必须大于0"
        return d, None
    except (InvalidOperation, ValueError):
        return None, f"{field_name}格式错误: {value}"


def _parse_date(value) -> tuple[date | None, str | None]:
    if value is None:
        return None, "日期不能为空"
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date(), None
    except ValueError:
        return None, f"日期格式错误: {value}（应为 YYYY-MM-DD）"


async def import_holdings(
    db: AsyncSession, file_content: bytes
) -> dict:
    """Import holdings from Excel. Returns success/error summary."""
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb["持仓导入"] if "持仓导入" in wb.sheetnames else wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    cols = _resolve_columns(
        header_row, HOLDING_FIELD_ALIASES, HOLDING_REQUIRED_FIELDS, LEGACY_HOLDING_COLS
    )

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    results = {"success": [], "errors": []}

    for i, row in enumerate(rows, start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        symbol_raw = _cell(row, cols, "symbol")
        name_raw = _cell(row, cols, "name")
        asset_type_raw = _cell(row, cols, "asset_type")
        account_raw = _cell(row, cols, "account")

        symbol = str(symbol_raw).strip() if symbol_raw else None
        name = str(name_raw).strip() if name_raw else None
        asset_type = str(asset_type_raw).strip() if asset_type_raw else None
        account = str(account_raw).strip() if account_raw else None

        errors = []
        if not symbol:
            errors.append("标识代码不能为空")
        if not name:
            errors.append("名称不能为空")
        if asset_type not in VALID_ASSET_TYPES:
            errors.append(f"资产类型无效: {asset_type}")

        quantity, err = _parse_decimal(_cell(row, cols, "quantity"), "数量")
        if err:
            errors.append(err)

        cost_cell = _cell(row, cols, "cost_price")
        if asset_type == "现金" and (
            cost_cell is None or (isinstance(cost_cell, str) and not cost_cell.strip())
        ):
            cost_price = Decimal("1")
        else:
            cost_price, err = _parse_decimal(cost_cell, "单位成本")
            if err:
                errors.append(err)

        latest_price = None
        latest_err = None
        if cols.get("latest_price") is not None:
            latest_price, latest_err = _parse_decimal_optional(
                _cell(row, cols, "latest_price"), "最新价"
            )
            if latest_err:
                errors.append(latest_err)

        if errors:
            results["errors"].append({"row": i, "error": "; ".join(errors)})
            continue

        holding = Holding(
            symbol=symbol,
            name=name,
            asset_type=asset_type,
            quantity=quantity,
            cost_price=cost_price,
            latest_price=latest_price,
            latest_price_updated_at=datetime.utcnow() if latest_price else None,
            account=account,
        )
        db.add(holding)
        results["success"].append({"row": i, "symbol": symbol, "name": name})

    if results["success"]:
        await db.commit()

    return results


async def import_transactions(
    db: AsyncSession, file_content: bytes, user_id
) -> dict:
    """Import transactions from Excel. Returns success/error summary."""
    from sqlalchemy import select

    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb["交易导入"] if "交易导入" in wb.sheetnames else wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    cols = _resolve_columns(
        header_row,
        TRANSACTION_FIELD_ALIASES,
        TRANSACTION_REQUIRED_FIELDS,
        LEGACY_TRANSACTION_COLS,
    )

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    results = {"success": [], "errors": []}

    for i, row in enumerate(rows, start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        symbol_raw = _cell(row, cols, "symbol")
        tx_type_raw = _cell(row, cols, "tx_type")
        symbol = str(symbol_raw).strip() if symbol_raw else None
        tx_type = str(tx_type_raw).strip() if tx_type_raw else None

        errors = []
        if not symbol:
            errors.append("标识代码不能为空")
        if tx_type not in VALID_TX_TYPES:
            errors.append(f"交易类型无效: {tx_type}")

        quantity, err = _parse_decimal(_cell(row, cols, "quantity"), "数量")
        if err:
            errors.append(err)
        price, err = _parse_decimal(_cell(row, cols, "price"), "价格")
        if err:
            errors.append(err)
        tx_date, err = _parse_date(_cell(row, cols, "tx_date"))
        if err:
            errors.append(err)

        fee = Decimal("0")
        fee_cell = _cell(row, cols, "fee")
        if fee_cell is not None and str(fee_cell).strip():
            fee_parsed, err = _parse_decimal(fee_cell, "手续费")
            if err:
                errors.append(err)
            else:
                fee = fee_parsed

        if errors:
            results["errors"].append({"row": i, "error": "; ".join(errors)})
            continue

        h_result = await db.execute(
            select(Holding).where(Holding.symbol == symbol)
        )
        holding = h_result.scalar_one_or_none()
        if not holding:
            results["errors"].append(
                {"row": i, "error": f"持仓不存在: {symbol}，请先添加持仓"}
            )
            continue

        transaction = Transaction(
            holding_id=holding.id,
            symbol=symbol,
            type=tx_type,
            quantity=quantity,
            price=price,
            fee=fee,
            date=tx_date,
            user_id=user_id,
        )
        db.add(transaction)
        results["success"].append({"row": i, "symbol": symbol, "type": tx_type})

    if results["success"]:
        await db.commit()

    return results
